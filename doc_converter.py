#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Office Belgelerini PDF'e Dönüştürme Betiği
Desteklenen formatlar: DOCX, XLSX, CSV, TXT
"""

import sys
import os
import base64
import json
import tempfile
import io
from fpdf import FPDF
import pandas as pd

# DOCX belgelerini işlemek için
try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

# Excel belgelerini işlemek için
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def docx_to_pdf(input_data):
    """DOCX dosyasını PDF'e dönüştürür"""
    # Geçici dosya oluştur
    with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as temp_docx:
        temp_docx.write(input_data)
        temp_docx_path = temp_docx.name

    try:
        # DOCX dosyasını oku
        document = Document(temp_docx_path)
        
        # PDF oluştur
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Başlık ve paragrafları ekle
        for paragraph in document.paragraphs:
            # Paragraf stiline göre fontları ayarla
            if paragraph.style.name.startswith('Heading'):
                pdf.set_font("Arial", 'B', size=14)  # Başlıklar kalın
            else:
                pdf.set_font("Arial", size=12)  # Normal metin
            
            # Metin ekle
            if paragraph.text.strip():  # Boş paragrafları atla
                # UTF-8 kodlamasıyla metni ekle
                text = paragraph.text.encode('latin-1', 'replace').decode('latin-1')
                pdf.multi_cell(0, 10, text)
        
        # PDF'i belleğe aktar
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        
        # Geçici dosyayı temizle
        os.unlink(temp_docx_path)
        
        return pdf_bytes
    except Exception as e:
        # Hata durumunda temizlik
        if os.path.exists(temp_docx_path):
            os.unlink(temp_docx_path)
        raise e


def xlsx_to_pdf(input_data):
    """Excel dosyasını PDF'e dönüştürür"""
    # Geçici dosya oluştur
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_xlsx:
        temp_xlsx.write(input_data)
        temp_xlsx_path = temp_xlsx.name
    
    try:
        # Excel dosyasını oku
        excel_data = pd.read_excel(temp_xlsx_path, sheet_name=None)
        
        # PDF oluştur
        pdf = FPDF()
        
        # Her çalışma sayfası için
        for sheet_name, df in excel_data.items():
            pdf.add_page()
            pdf.set_font("Arial", 'B', size=16)
            pdf.cell(0, 10, f"Çalışma Sayfası: {sheet_name}", ln=True)
            
            pdf.ln(5)
            pdf.set_font("Arial", 'B', size=12)
            
            # Sütun başlıklarını ekle
            col_width = pdf.w / len(df.columns)
            for col in df.columns:
                pdf.cell(col_width, 10, str(col), border=1)
            pdf.ln()
            
            # Verileri ekle
            pdf.set_font("Arial", size=10)
            for _, row in df.iterrows():
                for value in row:
                    # Değeri metin olarak dönüştür ve sığdır
                    value_str = str(value)
                    if len(value_str) > 15:  # Uzun içeriği kısalt
                        value_str = value_str[:12] + "..."
                    pdf.cell(col_width, 10, value_str, border=1)
                pdf.ln()
        
        # PDF'i belleğe aktar
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        
        # Geçici dosyayı temizle
        os.unlink(temp_xlsx_path)
        
        return pdf_bytes
    except Exception as e:
        # Hata durumunda temizlik
        if os.path.exists(temp_xlsx_path):
            os.unlink(temp_xlsx_path)
        raise e


def csv_to_pdf(input_data):
    """CSV dosyasını PDF'e dönüştürür"""
    try:
        # CSV dosyasını pandas ile oku
        csv_content = input_data.decode('utf-8')
        df = pd.read_csv(io.StringIO(csv_content))
        
        # PDF oluştur
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', size=16)
        pdf.cell(0, 10, "CSV Verileri", ln=True)
        
        pdf.ln(5)
        pdf.set_font("Arial", 'B', size=12)
        
        # Sütun başlıklarını ekle
        col_width = pdf.w / len(df.columns)
        for col in df.columns:
            pdf.cell(col_width, 10, str(col), border=1)
        pdf.ln()
        
        # Verileri ekle
        pdf.set_font("Arial", size=10)
        for _, row in df.iterrows():
            for value in row:
                # Değeri metin olarak dönüştür ve sığdır
                value_str = str(value)
                if len(value_str) > 15:  # Uzun içeriği kısalt
                    value_str = value_str[:12] + "..."
                pdf.cell(col_width, 10, value_str, border=1)
            pdf.ln()
        
        # PDF'i belleğe aktar
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        
        return pdf_bytes
    except Exception as e:
        raise e


def txt_to_pdf(input_data):
    """Metin dosyasını PDF'e dönüştürür"""
    try:
        # Metin içeriğini çıkar
        text_content = input_data.decode('utf-8')
        
        # PDF oluştur
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        # Metni satır satır ekle
        for line in text_content.split('\n'):
            # UTF-8 kodlamasıyla metni ekle
            text = line.encode('latin-1', 'replace').decode('latin-1')
            pdf.multi_cell(0, 8, text)
        
        # PDF'i belleğe aktar
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        
        return pdf_bytes
    except Exception as e:
        raise e


def html_to_pdf(input_data):
    """HTML dosyasını PDF'e dönüştürür"""
    try:
        # HTML içeriğini çıkar
        html_content = input_data.decode('utf-8')
        
        # PDF oluştur
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        # HTML başlığını bul
        import re
        title_match = re.search(r'<title>(.*?)</title>', html_content, re.IGNORECASE)
        if title_match:
            title = title_match.group(1)
            pdf.set_font("Arial", 'B', size=16)
            pdf.cell(0, 10, title, ln=True)
        
        # Basit bir HTML temizleme yapalım
        # (gerçek bir uygulama için daha kapsamlı bir HTML parser kullanılmalı)
        text_content = html_content
        # HTML etiketlerini kaldır
        text_content = re.sub(r'<[^>]*>', ' ', text_content)
        # Fazla boşlukları temizle
        text_content = re.sub(r'\s+', ' ', text_content).strip()
        
        # Metni ekle
        pdf.set_font("Arial", size=12)
        text = text_content.encode('latin-1', 'replace').decode('latin-1')
        pdf.multi_cell(0, 8, text)
        
        # PDF'i belleğe aktar
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        
        return pdf_bytes
    except Exception as e:
        raise e


def rtf_to_pdf(input_data):
    """RTF dosyasını PDF'e dönüştürür"""
    try:
        # RTF içeriği için basit bir çözüm
        # (daha iyi sonuçlar için özel RTF parser kullanılabilir)
        rtf_content = input_data.decode('utf-8', errors='ignore')
        
        # RTF etiketlerini kaldır (çok basit bir yaklaşım)
        import re
        # RTF başlığını kaldır
        text_content = re.sub(r'^.*?\\rtf1.*?{', '', rtf_content, flags=re.DOTALL)
        # Kontrol karakterlerini kaldır
        text_content = re.sub(r'\\[a-z0-9]+', ' ', text_content)
        # Kapalı süslü parantezleri kaldır
        text_content = re.sub(r'}', '', text_content)
        # Fazla boşlukları temizle
        text_content = re.sub(r'\s+', ' ', text_content).strip()
        
        # PDF oluştur
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        # Metni ekle
        text = text_content.encode('latin-1', 'replace').decode('latin-1')
        pdf.multi_cell(0, 8, text)
        
        # PDF'i belleğe aktar
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        
        return pdf_bytes
    except Exception as e:
        raise e


def convert_to_pdf(input_data, mime_type, file_name):
    """Belgeyi türüne göre PDF'e dönüştürür"""
    try:
        # Dosya türüne göre uygun dönüştürücüyü çağır
        if mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or file_name.endswith('.docx'):
            if not DOCX_AVAILABLE:
                raise ImportError("python-docx kütüphanesi yüklü değil")
            return docx_to_pdf(input_data)
            
        elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or file_name.endswith('.xlsx'):
            if not EXCEL_AVAILABLE:
                raise ImportError("openpyxl kütüphanesi yüklü değil")
            return xlsx_to_pdf(input_data)
            
        elif mime_type == "text/csv" or file_name.endswith('.csv'):
            return csv_to_pdf(input_data)
            
        elif mime_type == "text/plain" or file_name.endswith('.txt'):
            return txt_to_pdf(input_data)
            
        elif mime_type == "text/html" or file_name.endswith('.html') or file_name.endswith('.htm'):
            return html_to_pdf(input_data)
            
        elif mime_type == "application/rtf" or file_name.endswith('.rtf'):
            return rtf_to_pdf(input_data)
            
        else:
            raise ValueError(f"Desteklenmeyen dosya türü: {mime_type}")
    
    except Exception as e:
        error_message = f"Dönüştürme hatası: {str(e)}"
        sys.stderr.write(error_message)
        raise


def main():
    """
    Ana fonksiyon - Node.js'den çağrılacak
    
    Beklenen format:
    1. Argüman: Base64 formatında dosya içeriği
    2. Argüman: MIME türü
    3. Argüman: Dosya adı
    
    Çıktı:
    Base64 formatında PDF içeriği (stdout'a yazılır)
    """
    try:
        if len(sys.argv) < 4:
            raise ValueError("Eksik argümanlar. Beklenen format: <base64_file_content> <mime_type> <file_name>")
        
        # Argümanları al
        base64_content = sys.argv[1]
        mime_type = sys.argv[2]
        file_name = sys.argv[3]
        
        # Base64'ü decode et
        file_content = base64.b64decode(base64_content)
        
        # Dönüştürme işlemi
        pdf_bytes = convert_to_pdf(file_content, mime_type, file_name)
        
        # Sonucu base64 olarak encode et ve stdout'a yaz
        pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
        
        # JSON formatında sonuç döndür
        result = {
            "success": True,
            "pdf_base64": pdf_base64,
            "original_size": len(file_content),
            "pdf_size": len(pdf_bytes)
        }
        print(json.dumps(result))
        
    except Exception as e:
        # Hata durumunda hata mesajı döndür
        error_result = {
            "success": False,
            "error": str(e)
        }
        print(json.dumps(error_result))
        sys.exit(1)


if __name__ == "__main__":
    main()